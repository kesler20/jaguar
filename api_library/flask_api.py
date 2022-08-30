from flask import redirect, url_for, render_template, request, session, send_from_directory, flash
import os
import threading
import sqlalchemy
import datetime
import openpyxl as xl
from database_models import *

from datetime import timedelta
from flask import Flask, redirect, url_for, render_template, request, session, send_from_directory, flash
import os
from os import path as ps
import threading
from sqlalchemy import create_engine, Column, Integer, String, DateTime, ForeignKey
from sqlalchemy.orm import relationship
from flask_sqlalchemy import SQLAlchemy
import datetime
import openpyxl as xl
import selenium
from selenium import webdriver as wb
import time 
from tkinter import Tk
from tkinter import filedialog
import pandas as pd

create_engine("mysql+pymysql://user:pw@host/db", pool_pre_ping=True)
ROOT_DIR = os.path.dirname(os.getcwd())
app = Flask(
    __name__, 
    template_folder=ps.join(ROOT_DIR, 'expertsystem-gh-pages', 'templates'),
    static_folder=ps.join(ROOT_DIR, 'expertsystem-gh-pages', 'static')
)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///user.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'top secret!'
app.secret_key = 'password'
app.permanent_session_lifetime = timedelta(minutes=50)
db = SQLAlchemy(app)

#------------------------------ BACKEND FUNCTIONALITY-----------------------------------------------

def get_specific_file_extensions_in_cwd(file_extension):
    _dir = os.path.curdir
    files = os.listdir(f'{_dir}')
    specific_files = []
    for file in files:
        if file.endswith(f'{file_extension}'):
            specific_files.append(file)
    return specific_files

def create_session_data(filename, x, y, z):
    
    db.create_all()
    session_data = SessionDataModel(
        session_id=increment_id(), 
        user_id=2, 
        filename=filename,
        x_simulation=x, 
        y_simulation=y, 
        z_simulation=z
    )
    db.session.add(session_data)
    db.session.commit()
    return session_data

def update_session_data(index,filename, x, y, z):
    db.create_all()
    datasess = db.session.query(SessionDataModel).all()
    datasess[1].filename = filename
    if x == '':
        pass
    else:
        datasess[1].x_simulation = x
        datasess[1].y_simulation = y
        datasess[1].z_simulation = z
    print(datasess)
    db.session.commit()

def upload_file():

    current_working_directory = os.path.curdir
    root = Tk()
    root.title('Web Crawler GUI')
    root_height = 500
    root_width = 500
    root.geometry(f'{root_height}x{root_width}')
    root.iconbitmap(r'static\img\favicon.ico')
    root.filename = filedialog.asksaveasfilename(initialdir=f'{current_working_directory}', title='upload a file')
    df = pd.read_csv(r'{}'.format(root.filename))
    variables = df.columns 
    x_y_z = []
    for var in variables:
        if var.startswith('Unnamed'):
            pass
        elif var == 0:
            pass
        elif var.startswith('0'):
            pass
        else:
            x_y_z.append(var)
    print(x_y_z)
    create_session_data(r'{}'.format(root.filename),x=x_y_z[0],y=x_y_z[1],z=x_y_z[2])
    root.mainloop()

def reset(reset_all_users=False, reset_all_posts=False):
    if 'username' in session:
        print('          loading reset .....')
        db.create_all()
        all_users = db.session.query(UserAccount).all()
        all_posts = db.session.query(Post).all()
        if reset_all_users:
            for user in all_users:
                users_to_delete = UserAccount.query.filter_by(name=user.participant_id).first()
                db.session.delete(users_to_delete)
                db.session.commit()
        else:
            pass
        if reset_all_posts:
            for post in all_posts:
                posts_to_delete = users_to_delete = Post.query.filter_by(session_id=post.session_id).first()
                db.session.delete(posts_to_delete)
                db.session.commit()
        else:
            pass
        print(all_users)
        session.pop('username', None) 
        session.pop('Node', None) 
        session.pop('date', None) 
    else:
        pass  
    
def increment_id():
    db.create_all()
    users = db.session.query(UserAccount).all()
    return len(users) + 1

def initialize_web_app():
    try:
        x = ps.abspath('msedgedriver.exe')
        driver = wb.Edge(x)
        driver.get('http://127.0.0.1:5500/')
    except selenium.common.exceptions.WebDriverException:
        print(threading.active_count())
        time.sleep(200000)
        
def init_session_(number_of_attributes, attribute_values):
    for i in range(number_of_attributes):
        session[attribute_values[i]] = request.form[attribute_values[i]]     

def modify_workbook_in_session(node, participant_id, date=datetime.datetime.now(),):
    if session.permanent:
        try:
            wb = xl.load_workbook('Expert_System_Session.xlsx', False)
        except PermissionError:
            flash('Close the Expert System file for 5 seconds')
            time.sleep(4)
            wb = xl.load_workbook('Expert_System_Session.xlsx', False)

        date_cell = wb['User Information']['C8'] 
        node_cell = wb['User Information']['C9']
        participant_cell = wb['User Information']['C10'] 
        node_cell.value = node
        participant_cell.value = participant_id
        date_cell.value = date
        
        wb.save('Expert_System_Session.xlsx')
        os.system('start Expert_System_Session.xlsx')

def check_database_column(column_objct):
    db.create_all()
    datasess = db.session.query(column_objct).all()
    print(datasess)

def get_session_by_index(index):
    db.create_all()
    datasess = db.session.query(SessionDataModel).all()
    print(datasess)
    session_data = datasess[index]
    return session_data
# -------------------------------- DATABASE MODEL OBJECTS----------------------------------------------
class UserAccount(db.Model):
    __tablename__ = 'user_account'
    participant_id = Column(Integer, primary_key=True, nullable=False)
    username = Column(String(80), nullable=False, default=f'username{participant_id}', unique=True)
    post = relationship('Post', backref='author', lazy=True)
    session_data = relationship('SessionDataModel', backref='creator', lazy=True)


    def __repr__(self):
        return f'''
        UserAccount(
                username : {self.username},
                participant id : {self.participant_id},
                session activity: {self.post}
            )
        '''
    
class Post(db.Model):

    session_id = Column(Integer, primary_key=True, nullable=False)
    content = Column(String(80), nullable=False, default='start session')
    user_id = Column(Integer, ForeignKey('user_account.participant_id'), nullable=False)
    date = Column(DateTime, nullable=False, default=datetime.datetime.utcnow())

    def __repr__(self):
        return f'''
            Post(
                date : {self.date},
                session id : {self.session_id},
                content: {self.content},
                author : {self.user_id}
            )
        '''

class SessionDataModel(db.Model):

    session_id = Column(Integer, primary_key=True, nullable=False)
    filename = Column(String(80), nullable=False, default=f'{ROOT_DIR}')
    user_id = Column(Integer, ForeignKey('user_account.participant_id'), nullable=False)
    x_simulation = Column(String(80), nullable=False, default=f'{ROOT_DIR}')
    y_simulation = Column(String(80), nullable=False, default=f'{ROOT_DIR}')
    z_simulation = Column(String(80), nullable=False, default=f'{ROOT_DIR}')
    consequence = Column(String(80), nullable=False, default=f'{ROOT_DIR}')
    consequence_value = Column(Integer, nullable=False, default=0)
    event_tree_probability = Column(Integer, nullable=False, default=0)
    initiating_event = Column(String(80), nullable=False, default=f'{ROOT_DIR}')

    def __repr__(self):
        return f'''
            SessionDataModel(
                session id : {self.session_id},
                filename: {self.filename},
                author : {self.user_id},
                x : {self.x_simulation},
                y : {self.y_simulation},
                z : {self.z_simulation},
                consequence : {self.consequence},
                consequence value : {self.consequence_value},
                event tree probability : {self.event_tree_probability},
                initiating event : {self.initiating_event}
            )
        '''
db.create_all()
ts = threading.Thread(target=initialize_web_app)
ts.start()


#------------------------------------FLASK APPLICATION-------------------------------------------------
@app.route('/static/<path:filename>')
def serve_static(filename):
    global ROOT_DIR

    return send_from_directory(os.path.join(ROOT_DIR, 'static', 'js'),   filename)    
@app.route('/logout', methods=['POST', 'GET'])
def logout():
    if 'username' in session:
        flash('You have been logged out!!')
        reset()
    else:
        pass
    return render_template('index.html')

@app.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':

        db.create_all()
        session.permanent = True

        date = request.form['date']
        node = request.form['Node']
        participant_username = request.form['participants']

        #store it in the database if new
        try:
            found_user = UserAccount.query.filter_by(username=participant_username).first()    
        except sqlalchemy.exc.OperationalError:
            found_user = False
        if found_user:
            session['username'] = found_user.username 
            session['Node'] = request.form['Node']
            session['date'] = date
            session['participants'] = participant_username
        else:
            user = UserAccount(username=participant_username, participant_id=increment_id())
            print(user)
            post = Post(content=node, user_id=user.participant_id)#since the id is our primary key it will be assigned automatically
            print(post)
            db.session.add(user)
            db.session.add(post)
            db.session.commit()
            
        #store it on the session 
        if 'username' in session:
            pass
        else:
            session['date'] = request.form['date']
            session['username'] = request.form['participants']
            session['Node'] = request.form['Node']

        return render_template('login.html',todays_date=date, participants=participant_username)
    # in the event the request method is a get request
    else:
        if 'username' in session:
            date = session['date']
            participant_username = session['username']
            flash(f'Hello {participant_username}!!')
            return render_template('login.html', todays_date=date, participants=participant_username)
        else:
            flash(f'Hello please enter your details!!')
            return render_template('login.html', todays_date=datetime.datetime.now())
    

@app.route("/")
def render_home_page():
    return render_template('index.html')

@app.route('/simulation', methods=['POST', 'GET'])
def simulation():
    if request.method == 'POST':
        return redirect(url_for('data_analysis'))
    else:
        flash('Please enter data')
        return render_template('simulation.html')

@app.route("/<page>/")
def render_page(page):
    if page == 'home':
        return render_template('index.html')
    
    elif page == 'favicon.ico':
        pass

    elif page == 'index.html':
        print('start------------------------------')
        return render_template('index.html')

    elif page == 'ExpertSystem':
        return redirect(url_for('expert'))

    elif page == 'fault_tree':
        #import GUI_process_safety
        return render_template('simulation.html')

    elif page == 'web_crawler':
        import web_crawler
        return render_template('chemicals_edit.html')

    elif page == 'data_upload':
        upload_file()
        return render_template('simulation.html')

    else:
        return render_template(r'{}.html'.format(page))

@app.route('/console_database', methods=['POST', 'GET'])
def console_database():
    if request.method == 'POST':
        lel_kg = int(request.form['lel-kg'])
        lel_v = int(request.form['lel-v'])
        import test_ExpertSystem as ts
        lel = ts.lower_explosible_limit_calculations(lel_v,lel_kg)
        flash(f'Lower explosible limit: {lel}')

        #rho = request.form['rho']
        #epsilon0 = request.form['epsilon0']
        #epsionR = request.form['epsionR']
        #response = charge_dissipation(epsilon0, epsionR, rho)
        #flash(response)

        return render_template('console_database.html')
    else:
        flash('result: ')
        return render_template('console_database.html')

@app.route('/data_analysis', methods=['POST','GET'])
def data_analysis():
    if request.method == 'POST':
        filename = request.form['simulation-filename']
        X = request.form['simulation-X']
        Y = request.form['simulation-Y']
        Z = request.form['simulation-Z']
        
        all_data = db.session.query(SessionDataModel).all()
        sessdt = update_session_data(len(all_data)-1,filename, X, Y, Z)
        print(sessdt) 
        import simulation  
        return render_template('data_analysis.html')
    else:
        return render_template('data_analysis.html')

@app.route('/data_analysis2', methods=['POST','GET'])
def data_analysis2():
    
    if request.method == 'POST':
        import Event_tree_backend
        return render_template('data_analysis2.html')
    else:
        return render_template('data_analysis2.html')

@app.route("/ExpertSystem")
def expert():
    try:
        if 'username' in session:
            date = session['date']
            node = session['Node']
            participant_username = session['username']
        else:
            date = ''
            all_posts = db.session.query(Post).all()
            last_inspected_node = all_posts[len(all_posts)-1].content
            node = last_inspected_node
            participant_username = 'Please Enter'

        wb = xl.load_workbook('Expert_System.xlsm')
        ws = wb['User Information']
        date_cell = ws['C8']
        node_cell = ws['C9']
        participant_cell = ws['C10']

        if date == '':
            date_cell.value = datetime.datetime.now()
        else:
            date_cell.value = date
        node_cell.value = node
        participant_cell.value = participant_username
        wb.save('Expert_System_Session.xlsx')
        os.system('start Expert_System_Session.xlsx')
    except PermissionError:
        return '''
        <h3>Please close the Expert_System_Session.xlsx and Expert_System.xlsm files</h3>
        <h4> Then Logout and try the Login again</h4>
        '''
    except UserWarning:
        t = threading.Thread(target=initialize_web_app)
        t.start()
    
    os.system('start Expert_System_Session.xlsx')
    return redirect(url_for('render_home_page'))


#----------------------------------- COMMAND TO START-----------------------------------------------
if __name__ == '__main__':
    app.run(debug=True, port=5500)
     
    
    
